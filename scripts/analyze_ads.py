"""
BRAND 廣告效益分析腳本
用法：
  python analyze_ads.py --full 廣告報表.csv --7d 7天.csv --4d 4天.csv
                        --leads Systeme名單.csv --notion notion.json
                        --output 輸出.xlsx [--period "2026年4月W16上週"]
"""
import warnings; warnings.filterwarnings("ignore")
import csv, sys, argparse, json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CLI ────────────────────────────────────────────────────────────────────────
p = argparse.ArgumentParser()
p.add_argument('--full',      required=True)
p.add_argument('--7d',        required=True, dest='d7')
p.add_argument('--4d',        required=True, dest='d4')
p.add_argument('--leads',     required=True)
p.add_argument('--all-leads', default=None, dest='all_leads')
p.add_argument('--consult',   default=None)
p.add_argument('--notion',    default=None)
p.add_argument('--output',    required=True)
p.add_argument('--period',    default='')
p.add_argument('--cpl-threshold', type=float, default=None)
args = p.parse_args()

# ── Helpers ────────────────────────────────────────────────────────────────────
FONT = "Arial"

def load_csv(path, enc="utf-8"):
    try:
        with open(path, encoding=enc) as f:
            return list(csv.DictReader(f))
    except UnicodeDecodeError:
        with open(path, encoding="utf-16") as f:
            return list(csv.DictReader(f))

def sf(v, d=None):
    try: return float(v)
    except: return d

def si(v, d=0):
    try: return int(v)
    except: return d

thin = Side(style="thin", color="CCCCCC")
def tb(): return Border(left=thin, right=thin, top=thin, bottom=thin)
def fl(h): return PatternFill("solid", start_color=h, fgColor=h)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def no_wrap(align="left"): return Alignment(horizontal=align, vertical="center", wrap_text=False)
def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def wc(ws, row, col, val, bold=False, bg=None, nf=None, align=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=FONT, bold=bold, size=10, color=color)
    if bg: c.fill = fl(bg)
    if nf: c.number_format = nf
    c.alignment = align or left()
    c.border = tb()
    return c

def hdr(ws, row, cols, bg="2C3E50"):
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        c.fill = fl(bg); c.alignment = center(); c.border = tb()
    ws.row_dimensions[row].height = 36

def section_title(ws, row, text, n_cols, bg="1A2E44"):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    c.fill = fl(bg); c.alignment = left()
    ws.row_dimensions[row].height = 26

# ── Load data ──────────────────────────────────────────────────────────────────
ads_full    = load_csv(args.full)
ads_7d      = load_csv(args.d7)
ads_4d      = load_csv(args.d4)
systeme     = load_csv(args.leads)
systeme_all = load_csv(args.all_leads) if args.all_leads else systeme

consults = []
if args.consult:
    consults = load_csv(args.consult)

notion_data = []
if args.notion:
    with open(args.notion) as f:
        notion_data = json.load(f)

# ── Build lookup structures ────────────────────────────────────────────────────
email_to_s = {}
for r in sorted(systeme_all, key=lambda x: x.get('Date Registered', '')):
    e = r['电子邮件'].strip().lower()
    if e not in email_to_s:
        email_to_s[e] = r

leads_by_utm = defaultdict(int)
for r in systeme:
    leads_by_utm[r.get('utm_content','').strip()] += 1

brand_consults = []
if notion_data:
    for n in notion_data:
        brand_consults.append({
            'email': n.get('email','').strip().lower(),
            'name':  n.get('name',''),
            'status': n.get('status',''),
            'date':  n.get('date',''),
            'source': 'notion'
        })
elif consults:
    for c in consults:
        if 'B.R.A.N.D' in c.get('Event Type Name',''):
            brand_consults.append({
                'email': c['Invitee Email'].strip().lower(),
                'name':  c['Invitee Name'],
                'status': '',
                'date':  c.get('Start Date & Time',''),
                'source': 'calendly'
            })

consult_by_utm = defaultdict(list)
for c in brand_consults:
    s = email_to_s.get(c['email'])
    if s:
        utm = s.get('utm_content','').strip()
        if utm and utm != 'link_in_bio':
            consult_by_utm[utm].append(c)

def has_roi(utm):
    return len(consult_by_utm.get(utm, [])) > 0

# ── CPL thresholds ─────────────────────────────────────────────────────────────
CPL_GOOD = args.cpl_threshold if args.cpl_threshold else 5.0
CPL_WARN = 8.0

ad_7d_map = {a['廣告編號']: a for a in ads_7d}
ad_4d_map = {a['廣告編號']: a for a in ads_4d}

def cpl_tier(cpl):
    if cpl is None: return None
    if cpl <= CPL_GOOD: return 'good'
    if cpl < CPL_WARN:  return 'warn'
    return 'bad'

# ── Build per-ad analysis ──────────────────────────────────────────────────────
def judge(ad_id, cpl_full):
    tier     = cpl_tier(cpl_full)
    roi_high = has_roi(ad_id)
    a7 = ad_7d_map.get(ad_id, {})
    a4 = ad_4d_map.get(ad_id, {})
    cpl_7d = sf(a7.get('每次成果成本'))
    cpl_4d = sf(a4.get('每次成果成本'))

    if   tier == 'good' and roi_high:      quad = "⭐ 優等廣告"
    elif tier == 'warn' and roi_high:      quad = "🔵 第二等"
    elif tier == 'bad'  and roi_high:      quad = "🟠 高CPL有轉換"
    elif tier == 'good' and not roi_high:  quad = "⚠️ 第三等"
    elif tier == 'warn' and not roi_high:  quad = "🟡 警示觀察"
    else:                                  quad = "🚫 垃圾廣告"

    t4, t7 = cpl_tier(cpl_4d), cpl_tier(cpl_7d)
    d4g, d7g = t4 == 'good', t7 == 'good'
    if cpl_4d is None or cpl_7d is None:
        trend = "⚪ 資料不足"; trend_act = "待觀察"
    elif d4g and d7g:         trend = "📈 4天好＋7天好"; trend_act = "加預算"
    elif not d4g and not d7g: trend = "📉 4天差＋7天差"; trend_act = "立即關閉"
    elif d4g and not d7g:     trend = "📊 4天好＋7天差"; trend_act = "繼續觀察"
    else:                     trend = "🔄 4天差＋7天好"; trend_act = "觀察到下階段"

    n = len(consult_by_utm.get(ad_id, []))
    if n > 0 and trend_act == "立即關閉":
        final = "⚠️ 曾帶來諮詢但CPL惡化，建議暫停優化"
    elif quad == "⭐ 優等廣告" and trend_act == "加預算": final = "✅ 強烈建議加碼"
    elif quad == "⭐ 優等廣告":                           final = "👀 優質素材，持續監控趨勢"
    elif quad == "🚫 垃圾廣告" and trend_act == "立即關閉": final = "❌ 立即關閉"
    elif quad == "🚫 垃圾廣告":                           final = "❌ 建議關閉（無諮詢轉換）"
    elif quad == "⚠️ 第三等":                             final = "⚠️ 謹慎保留，觀察是否帶出諮詢"
    elif quad == "🟡 警示觀察":                           final = "🟡 CPL偏高且無轉換，密切監控"
    elif quad == "🔵 第二等":                             final = "🔵 保留觀察（CPL偏高但有轉換潛力）"
    elif quad == "🟠 高CPL有轉換":                        final = "🟠 CPL過高，評估是否優化素材"
    else:                                                final = "⚪ 待觀察"

    return dict(quad=quad, trend=trend, trend_act=trend_act, final=final,
                cpl_7d=cpl_7d, cpl_4d=cpl_4d,
                d4_str="好" if d4g else ("差" if cpl_4d else "—"),
                d7_str="好" if d7g else ("差" if cpl_7d else "—"))

rows = []
for a in ads_full:
    ad_id  = a['廣告編號']
    j      = judge(ad_id, sf(a.get('每次成果成本')))
    a7     = ad_7d_map.get(ad_id, {})
    a4     = ad_4d_map.get(ad_id, {})
    c_list = consult_by_utm.get(ad_id, [])
    j_clean = {k: v for k, v in j.items() if k not in ['cpl_7d', 'cpl_4d']}
    rows.append(dict(
        ad_id=ad_id, ad_name=a['廣告名稱'], ad_group=a.get('廣告組合名稱',''),
        ad_status=a.get('廣告投遞',''),
        cpl_full=sf(a.get('每次成果成本')), spend_full=sf(a.get('花費金額 (USD)')),
        leads_full=si(a.get('成果')),
        reach=si(a.get('觸及人數')), impressions=si(a.get('曝光次數')),
        cpl_7d=j['cpl_7d'], spend_7d=sf(a7.get('花費金額 (USD)')), leads_7d=si(a7.get('成果')),
        cpl_4d=j['cpl_4d'], spend_4d=sf(a4.get('花費金額 (USD)')), leads_4d=si(a4.get('成果')),
        s_leads=leads_by_utm.get(ad_id, 0),
        n_consult=len(c_list),
        n_converted=sum(1 for c in c_list if c.get('status') == '成交'),
        n_followup=sum(1 for c in c_list if c.get('status') == '需跟進'),
        n_noshow=sum(1 for c in c_list if c.get('status') == '未出席'),
        cpql=sf(a.get('花費金額 (USD)'))/len(c_list) if len(c_list) > 0 else None,
        cpql_attended=(sf(a.get('花費金額 (USD)'))/(len(c_list)-sum(1 for c in c_list if c.get('status')=='未出席'))
                       if (len(c_list)-sum(1 for c in c_list if c.get('status')=='未出席')) > 0 else None),
        consult_names="、".join(c['name'] for c in c_list) or "—",
        **j_clean
    ))

# ── Consult source breakdown ────────────────────────────────────────────────────
ad_name_map  = {a['廣告編號']: a['廣告名稱'] for a in ads_full}
ad_spend_map = {a['廣告編號']: sf(a.get('花費金額 (USD)')) for a in ads_full}

cat_colors = {
    "✅ Meta廣告（可追蹤）": ("D4EDDA","1E7E34"),
    "🔗 Link in Bio":        ("FFF3CD","856404"),
    "❓ 無廣告編號（其他）":  ("FEF9E7","5A4E00"),
    "❌ 不在Systeme名單":     ("F8D7DA","721C24"),
}
detail_rows = []
for c in brand_consults:
    e   = c['email']
    s   = email_to_s.get(e)
    utm = s.get('utm_content','').strip() if s else None
    ad_n  = ad_name_map.get(utm,'') if utm else ''
    ad_sp = ad_spend_map.get(utm) if utm else None
    if not s:
        cat = "❌ 不在Systeme名單"
    elif utm and utm != 'link_in_bio' and ad_n:
        cat = "✅ Meta廣告（可追蹤）"
    elif utm == 'link_in_bio':
        cat = "🔗 Link in Bio"
    else:
        cat = "❓ 無廣告編號（其他）"
    detail_rows.append((c, utm, ad_n, ad_sp, cat))

sort_order = {"✅ Meta廣告（可追蹤）":0,"🔗 Link in Bio":1,"❓ 無廣告編號（其他）":2,"❌ 不在Systeme名單":3}
detail_rows.sort(key=lambda x: sort_order.get(x[4], 9))

cat_counts = defaultdict(int)
for _, _, _, _, cat in detail_rows:
    cat_counts[cat] += 1

# ── Stage 2 check ──────────────────────────────────────────────────────────────
STAGE2_MIN_CONSULTS = 5
STAGE2_NOSHOW_THR   = 0.4
STAGE2_QUALITY_THR  = 0.2

stage2_alerts = []
for r in rows:
    if r['n_consult'] < STAGE2_MIN_CONSULTS:
        continue
    noshow_rate   = r['n_noshow'] / r['n_consult']
    quality_rate  = (r['n_converted'] + r['n_followup']) / r['n_consult']
    triggered = []
    if noshow_rate >= STAGE2_NOSHOW_THR:
        triggered.append(f"未出席率 {noshow_rate:.0%}（≥{STAGE2_NOSHOW_THR:.0%}）")
    if quality_rate <= STAGE2_QUALITY_THR:
        triggered.append(f"有效諮詢率 {quality_rate:.0%}（≤{STAGE2_QUALITY_THR:.0%}）")
    if triggered:
        stage2_alerts.append((r['ad_name'], r['n_consult'], noshow_rate, quality_rate, "、".join(triggered)))

# ── Color maps ─────────────────────────────────────────────────────────────────
QUAD_C = {
    "⭐ 優等廣告":    ("D4EDDA","1E7E34"),
    "🔵 第二等":      ("D1ECF1","0C5460"),
    "🟠 高CPL有轉換": ("FFE8CC","7A3B00"),
    "⚠️ 第三等":      ("FFF3CD","856404"),
    "🟡 警示觀察":    ("FFFACD","6B5900"),
    "🚫 垃圾廣告":    ("F8D7DA","721C24"),
}
TREND_C = {
    "📈 4天好＋7天好": ("D4EDDA","1E7E34"),
    "📊 4天好＋7天差": ("FFF3CD","856404"),
    "🔄 4天差＋7天好": ("D1ECF1","0C5460"),
    "📉 4天差＋7天差": ("F8D7DA","721C24"),
    "⚪ 資料不足":     ("F8F9FA","6C757D"),
}
FINAL_C = {
    "✅ 強烈建議加碼":                       ("D4EDDA","155724"),
    "👀 優質素材，持續監控趨勢":             ("D1ECF1","0C5460"),
    "⚠️ 曾帶來諮詢但CPL惡化，建議暫停優化": ("FFF3CD","856404"),
    "⚠️ 謹慎保留，觀察是否帶出諮詢":        ("FFF3CD","856404"),
    "🟡 CPL偏高且無轉換，密切監控":          ("FFFACD","6B5900"),
    "🔵 保留觀察（CPL偏高但有轉換潛力）":    ("D1ECF1","0C5460"),
    "🟠 CPL過高，評估是否優化素材":          ("FFE8CC","7A3B00"),
    "❌ 立即關閉":                           ("F8D7DA","721C24"),
    "❌ 建議關閉（無諮詢轉換）":             ("F8D7DA","721C24"),
    "⚪ 待觀察":                             ("F8F9FA","6C757D"),
}

# ── Summary stats ──────────────────────────────────────────────────────────────
total_spend      = sum(r['spend_full'] or 0 for r in rows)
total_meta_leads = sum(r['leads_full'] for r in rows)
total_sys_leads  = len(systeme)
total_consults   = len(brand_consults)
trackable_consults = cat_counts.get("✅ Meta廣告（可追蹤）", 0)

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1：📊 執行摘要
# ══════════════════════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "📊 執行摘要"
ws0.sheet_view.showGridLines = False

ws0.merge_cells("A1:F1")
ws0["A1"].value = f"BRAND 課程廣告效益分析報告｜{args.period}"
ws0["A1"].font  = Font(name=FONT, bold=True, size=14, color="FFFFFF")
ws0["A1"].fill  = fl("1A2E44"); ws0["A1"].alignment = center()
ws0.row_dimensions[1].height = 38

ws0.merge_cells("A2:F2")
ws0["A2"].value = f"數據來源：Meta 廣告報表 × Systeme 名單 × Notion 諮詢記錄｜CPL 判準：≤${CPL_GOOD:.0f} 好 / ${CPL_GOOD:.0f}–${CPL_WARN:.0f} 警示 / ≥${CPL_WARN:.0f} 不行"
ws0["A2"].font  = Font(name=FONT, size=9, color="555555")
ws0["A2"].alignment = center()
ws0.row_dimensions[2].height = 18

# 關鍵指標
section_title(ws0, 3, "▌ 關鍵指標總覽", 6)
overall_cpql = total_spend / trackable_consults if trackable_consults else None
total_noshow = sum(1 for c in brand_consults if c.get('status') == '未出席')
attended_consults = trackable_consults - sum(
    1 for c in brand_consults if c.get('status') == '未出席'
    and email_to_s.get(c['email']) and email_to_s[c['email']].get('utm_content','').strip()
    and email_to_s[c['email']].get('utm_content','').strip() != 'link_in_bio'
)
overall_cpql_attended = total_spend / attended_consults if attended_consults > 0 else None
metrics = [
    ("廣告總花費",                    f"${total_spend:,.2f} USD"),
    ("Meta 廣告帶來名單數（CPL）",    f"{total_meta_leads:,} 人｜${total_spend/total_meta_leads:.2f}/筆" if total_meta_leads else f"{total_meta_leads:,} 人"),
    ("Systeme 實際收到名單數",        f"{total_sys_leads:,} 人"),
    ("諮詢預約總數",                  f"{total_consults} 筆（未出席 {total_noshow} 筆）"),
    ("CPQL（含未出席）",              f"{trackable_consults} 筆｜${overall_cpql:.2f}/筆" if overall_cpql else f"{trackable_consults} 筆｜無歸因數據"),
    ("CP合格QL（排除未出席）",        f"{attended_consults} 筆｜${overall_cpql_attended:.2f}/筆" if overall_cpql_attended else f"{attended_consults} 筆｜無歸因數據"),
]
for i, (label, val) in enumerate(metrics, 4):
    ws0.row_dimensions[i].height = 22
    wc(ws0, i, 1, label, bold=True, bg="F0F4F8")
    ws0.merge_cells(f"B{i}:C{i}")
    wc(ws0, i, 2, val, bold=True, bg="FFFFFF", color="1A2E44")
    for col in range(4, 7):
        wc(ws0, i, col, "", bg="FFFFFF")

# 指標說明
note_row = len(metrics) + 4
ws0.row_dimensions[note_row].height = 14
notes = [
    ("CPQL（含未出席）",     "花費 ÷ 總諮詢數。反映廣告帶來諮詢的整體成本，含動機不足者。"),
    ("CP合格QL（排除未出席）","花費 ÷ 有出席諮詢數。排除未到場者，更準確反映廣告帶到有意願受眾的成本。"),
]
for i, (label, desc) in enumerate(notes):
    ri = note_row + i
    ws0.row_dimensions[ri].height = 20
    wc(ws0, ri, 1, label, bold=True, bg="EBF3FB", color="1A2E44")
    ws0.merge_cells(f"B{ri}:F{ri}")
    wc(ws0, ri, 2, desc, bg="EBF3FB", color="555555")

# 諮詢來源拆解
r_start = len(metrics) + len(notes) + 5
section_title(ws0, r_start, f"▌ 諮詢來源拆解（共 {total_consults} 筆）", 6)
hdr(ws0, r_start+1, ["來源類型","筆數","說明","","",""], bg="34495E")
source_desc = {
    "✅ Meta廣告（可追蹤）": "來自具體廣告編號，可對應廣告素材",
    "🔗 Link in Bio":        "從個人主頁連結進來，非廣告投放",
    "❓ 無廣告編號（其他）":  "在 Systeme 名單中但無廣告編號，可能為直接搜尋或分享",
    "❌ 不在Systeme名單":     "未透過名單磁鐵流程，可能直接看到諮詢連結",
}
for i, (cat, desc) in enumerate(source_desc.items()):
    ri = r_start + 2 + i
    ws0.row_dimensions[ri].height = 22
    bg, fg = cat_colors.get(cat, ("FFFFFF","000000"))
    wc(ws0, ri, 1, cat,  bold=True, bg=bg, color=fg)
    wc(ws0, ri, 2, cat_counts.get(cat, 0), bg=bg, color=fg, nf="#,##0", align=center())
    ws0.merge_cells(f"C{ri}:F{ri}")
    wc(ws0, ri, 3, desc, bg="FFFFFF")

# 未出席警示（唯一保留的品質信號）
r2 = r_start + 2 + len(source_desc) + 1
noshow_ads = [(r['ad_name'], r['n_consult'], r['n_noshow'], r['n_noshow']/r['n_consult'])
              for r in rows if r['n_consult'] > 0 and r['n_noshow']/r['n_consult'] >= 0.5]
if noshow_ads:
    section_title(ws0, r2, "▌ 出席率警示（未出席率 ≥ 50%）", 6, bg="856404")
    hdr(ws0, r2+1, ["廣告名稱","諮詢數","未出席數","未出席率","說明",""], bg="856404")
    for i, (name, nc, nn, nr) in enumerate(noshow_ads):
        ri = r2 + 2 + i
        ws0.row_dimensions[ri].height = 22
        wc(ws0, ri, 1, name, bg="FFF3CD")
        wc(ws0, ri, 2, nc,   bg="FFF3CD", nf="#,##0", align=center())
        wc(ws0, ri, 3, nn,   bg="FFF3CD", nf="#,##0", align=center())
        wc(ws0, ri, 4, nr,   bg="FFF3CD", nf="0%",    align=center(), bold=True, color="856404")
        ws0.merge_cells(f"E{ri}:F{ri}")
        wc(ws0, ri, 5, "廣告文案與課程期待可能有落差，建議檢視素材", bg="FFF3CD", color="856404")
    r2 = r2 + 2 + len(noshow_ads) + 1
else:
    r2 += 1

# 分析結論
section_title(ws0, r2, "▌ 分析結論", 6)
final_counts = defaultdict(int)
for r in rows:
    final_counts[r['final']] += 1

conclusions = []
strong = final_counts.get("✅ 強烈建議加碼", 0)
watch  = final_counts.get("👀 優質素材，持續監控趨勢", 0)
close  = sum(v for k, v in final_counts.items() if "關閉" in k)
warn   = sum(v for k, v in final_counts.items() if "謹慎" in k or "警示" in k or "偏高" in k)
special = final_counts.get("⚠️ 曾帶來諮詢但CPL惡化，建議暫停優化", 0)

if strong:   conclusions.append(f"【強烈加碼】{strong} 支廣告 CPL 優異且已帶進諮詢，建議立即增加預算。")
if watch:    conclusions.append(f"【優質監控】{watch} 支廣告表現良好，持續追蹤趨勢變化。")
if special:  conclusions.append(f"【特殊案例】{special} 支廣告曾帶來諮詢但近期 CPL 惡化，建議暫停優化後再決定去留。")
if close:    conclusions.append(f"【建議關閉】{close} 支廣告 CPL 超標且無諮詢轉換，建議停止投放，釋放預算給優質素材。")
if warn:     conclusions.append(f"【謹慎觀察】{warn} 支廣告 CPL 尚可但尚未帶出諮詢，繼續觀察，不急著加碼或砍掉。")
if trackable_consults == 0 and total_consults > 0:
    conclusions.append(f"【追蹤缺口】本期 {total_consults} 筆諮詢中，0 筆可歸因到廣告編號，建議確認落地頁廣告編號參數是否正確帶入。")

for i, text in enumerate(conclusions):
    ri = r2 + 1 + i
    ws0.row_dimensions[ri].height = 22
    ws0.merge_cells(f"A{ri}:F{ri}")
    wc(ws0, ri, 1, text, bg="F8F9FA")

for ci, w in enumerate([28,20,10,24,10,8], 1): cw(ws0, ci, w)
ws0.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2：📢 廣告素材效益
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("📢 廣告素材效益")
ws1.sheet_view.showGridLines = False

ws1.merge_cells("A1:R1")
ws1["A1"].value = f"廣告素材效益分析｜{args.period}｜CPL 判準：≤${CPL_GOOD:.0f} 好 / ${CPL_GOOD:.0f}–${CPL_WARN:.0f} 警示 / ≥${CPL_WARN:.0f} 不行"
ws1["A1"].font  = Font(name=FONT, bold=True, size=13, color="FFFFFF")
ws1["A1"].fill  = fl("1A2E44"); ws1["A1"].alignment = center()
ws1.row_dimensions[1].height = 34

use_notion = bool(notion_data)
hdrs1 = ["廣告名稱","廣告編號","廣告組合","狀態",
         "本期\nMeta成果","本期\nCPL","Systeme\n名單數","Systeme\nCPL",
         "本期花費\n(USD)","觸及\n人數","曝光\n次數",
         "諮詢\n數量","CPQL\n(USD)","CP合格QL\n(USD)","未出席\n數","諮詢者姓名",
         "四象限","趨勢\n(4/7天)","趨勢\n建議","最終建議"]
if use_notion:
    hdrs1 = hdrs1[:12] + ["成交\n數量","需跟進\n數量"] + hdrs1[12:]

hdr(ws1, 2, hdrs1)
total_row = len(rows) + 3

for ri, r in enumerate(rows, 3):
    ws1.row_dimensions[ri].height = 22
    is_alt = ri % 2 == 0
    base_bg = "EBF3FB" if is_alt else "FFFFFF"
    qbg, qfg = QUAD_C.get(r['quad'], (base_bg,"000000"))
    tbg, tfg = TREND_C.get(r['trend'], (base_bg,"000000"))
    fbg, ffg = FINAL_C.get(r['final'], (base_bg,"000000"))
    sys_cpl  = r['spend_full']/r['s_leads'] if r['s_leads'] else None
    base_vals = [
        r['ad_name'], r['ad_id'], r['ad_group'], r['ad_status'],
        r['leads_full'], r['cpl_full'],
        r['s_leads'] if r['s_leads'] else "—", sys_cpl,
        r['spend_full'], r['reach'], r['impressions'],
        r['n_consult'], r['cpql'], r['cpql_attended'], r['n_noshow'], r['consult_names'],
    ]
    base_nfs = [None,None,None,None,"#,##0","$#,##0.00","#,##0","$#,##0.00",
                "$#,##0.00","#,##0","#,##0","#,##0","$#,##0.00","$#,##0.00","#,##0",None]
    # col index → use center no-wrap for numeric/status cols
    num_cols = {5,6,7,8,9,10,11,12,13,14,15}
    if use_notion:
        base_vals = base_vals[:12] + [r['n_converted'], r['n_followup']] + base_vals[12:]
        base_nfs  = base_nfs[:12]  + ["#,##0","#,##0"] + base_nfs[12:]
        num_cols  = {5,6,7,8,9,10,11,12,13,14,15,16,17}
    decision_vals = [r['quad'], r['trend'], r['trend_act'], r['final']]
    decision_bgs  = [qbg, tbg, tbg, fbg]
    decision_fgs  = [qfg, tfg, tfg, ffg]
    all_vals = base_vals + decision_vals
    all_nfs  = base_nfs  + [None,None,None,None]
    n_base   = len(base_vals)
    for ci, (v, nf) in enumerate(zip(all_vals, all_nfs), 1):
        di = ci - n_base - 1
        if di >= 0:
            wc(ws1, ri, ci, v, bold=True, bg=decision_bgs[di], nf=nf,
               color=decision_fgs[di], align=center())
        else:
            al = no_wrap("center") if ci in num_cols else left()
            wc(ws1, ri, ci, v, bg=base_bg, nf=nf, align=al)

ws1.row_dimensions[total_row].height = 24
wc(ws1, total_row, 1, "合計", bold=True, bg="2C3E50", color="FFFFFF")
for ci in range(2, len(hdrs1)+1):
    wc(ws1, total_row, ci, "", bg="2C3E50")
meta_col = 5; spend_col = 9; consult_col = 12
ws1.cell(total_row, meta_col).value  = f"=SUM({get_column_letter(meta_col)}3:{get_column_letter(meta_col)}{total_row-1})"
ws1.cell(total_row, spend_col).value = f"=SUM({get_column_letter(spend_col)}3:{get_column_letter(spend_col)}{total_row-1})"
ws1.cell(total_row, consult_col).value = f"=SUM({get_column_letter(consult_col)}3:{get_column_letter(consult_col)}{total_row-1})"
for ci in [meta_col, spend_col, consult_col]:
    ws1.cell(total_row, ci).font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    ws1.cell(total_row, ci).number_format = "$#,##0.00" if ci == spend_col else "#,##0"

col_widths1 = [30,18,20,8,9,11,9,11,11,9,10,7,11,11,7,22,16,20,12,28]
if use_notion: col_widths1 = col_widths1[:12] + [7,7] + col_widths1[12:]
for ci, w in enumerate(col_widths1, 1): cw(ws1, ci, w)
ws1.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3：📋 諮詢比對明細
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📋 諮詢比對明細")
ws2.sheet_view.showGridLines = False
ws2.merge_cells("A1:H1")
ws2["A1"].value = f"BRAND 諮詢來源比對明細｜{args.period}"
ws2["A1"].font  = Font(name=FONT, bold=True, size=13, color="FFFFFF")
ws2["A1"].fill  = fl("1A2E44"); ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 34

h2 = ["諮詢時間","姓名","Email","在Systeme？","廣告編號","廣告名稱","廣告花費(USD)"]
if use_notion: h2 += ["成交狀態"]
h2 += ["來源分類"]
hdr(ws2, 2, h2)

for ri, (c, utm, ad_n, ad_sp, cat) in enumerate(detail_rows, 3):
    ws2.row_dimensions[ri].height = 22
    bg, fg = cat_colors.get(cat, ("FFFFFF","000000"))
    vals = [c.get('date',''), c['name'], c['email'],
            "是" if email_to_s.get(c['email']) else "否",
            utm or "（無）", ad_n or "（無）", ad_sp]
    if use_notion: vals.append(c.get('status','—'))
    vals.append(cat)
    nfs = [None,None,None,None,None,None,"$#,##0.00"] + ([None] if use_notion else []) + [None]
    for ci, (v, nf) in enumerate(zip(vals, nfs), 1):
        is_last = ci == len(vals)
        wc(ws2, ri, ci, v, bold=is_last, bg=bg, nf=nf,
           color=fg if is_last else "000000")

for ci, w in enumerate([16,14,28,7,20,24,11]+([11] if use_notion else [])+[22], 1):
    cw(ws2, ci, w)
ws2.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4：📥 Systeme 名單統計
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("📥 Systeme名單統計")
ws3.sheet_view.showGridLines = False
ws3.merge_cells("A1:G1")
ws3["A1"].value = f"Systeme 名單廣告編號統計｜{args.period}"
ws3["A1"].font  = Font(name=FONT, bold=True, size=13, color="FFFFFF")
ws3["A1"].fill  = fl("1A2E44"); ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 34

hdr(ws3, 2, ["廣告編號","廣告名稱","廣告花費(USD)","Meta成果數","Systeme名單數","諮詢數","名單→諮詢率"])

all_utms = sorted(leads_by_utm.keys(), key=lambda x: -leads_by_utm[x])
for ri, utm in enumerate(all_utms, 3):
    ws3.row_dimensions[ri].height = 22
    ad    = next((a for a in ads_full if a['廣告編號'] == utm), None)
    count = leads_by_utm[utm]
    n_c   = len(consult_by_utm.get(utm, []))
    bg    = "D4EDDA" if n_c > 0 else ("EBF3FB" if ri%2==0 else "FFFFFF")
    if utm == 'link_in_bio':
        an, asp, ml = "（Link in Bio）", None, None
    elif ad:
        an, asp, ml = ad['廣告名稱'], sf(ad.get('花費金額 (USD)')), si(ad.get('成果'))
    else:
        an, asp, ml = "（廣告報表中無對應）", None, None
    rate = n_c/count if count else 0
    vals = [utm or "（空白）", an, asp, ml, count, n_c, rate]
    nfs  = [None,None,"$#,##0.00","#,##0","#,##0","#,##0","0.0%"]
    for ci, (v, nf) in enumerate(zip(vals, nfs), 1):
        wc(ws3, ri, ci, v, bold=(ci==6 and n_c>0), bg=bg, nf=nf,
           color="1E7E34" if (ci==6 and n_c>0) else "000000")

for ci, w in enumerate([22,26,11,10,10,7,11], 1): cw(ws3, ci, w)
ws3.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5：🎯 廣告決策矩陣
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("🎯 廣告決策矩陣")
ws4.sheet_view.showGridLines = False
ws4.merge_cells("A1:O1")
ws4["A1"].value = f"廣告素材決策分析｜CPL × 諮詢率四象限 + 4天/7天趨勢矩陣｜{args.period}"
ws4["A1"].font  = Font(name=FONT, bold=True, size=13, color="FFFFFF")
ws4["A1"].fill  = fl("1A2E44"); ws4["A1"].alignment = center()
ws4.row_dimensions[1].height = 34

# PART 1：四象限
section_title(ws4, 2, "▌ PART 1｜CPL × 諮詢率四象限判斷", 15)
h4a = ["廣告名稱","廣告編號","本期花費\n(USD)","本期\nMeta成果","本期\nCPL",
       "Systeme\n名單數","諮詢數","CPQL\n(USD)","CP合格QL\n(USD)","名單→諮詢率","CPL\n判定","ROI\n判定","四象限\n分類","建議動作"]
hdr(ws4, 3, h4a)

for ri, r in enumerate(rows, 4):
    ws4.row_dimensions[ri].height = 22
    tier     = cpl_tier(r['cpl_full'])
    roi_high = r['n_consult'] > 0
    cpl_label = {"good":"低 ✓","warn":"中 △","bad":"高 ✗"}.get(tier,"—")
    roi_label = "高 ✓" if roi_high else "低 ✗"
    cpl_label_bg = {"good":"D4EDDA","warn":"FFF3CD","bad":"F8D7DA"}.get(tier,"FFFFFF")
    cpl_label_fg = {"good":"1E7E34","warn":"856404","bad":"721C24"}.get(tier,"000000")
    qbg, qfg = QUAD_C.get(r['quad'],("FFFFFF","000000"))
    action_map = {
        "✅ 強烈建議加碼":            "加碼投放",
        "👀 優質素材，持續監控趨勢":  "持續監控",
        "⚠️ 曾帶來諮詢但CPL惡化，建議暫停優化": "暫停優化",
        "⚠️ 謹慎保留，觀察是否帶出諮詢": "謹慎保留",
        "🟡 CPL偏高且無轉換，密切監控": "密切監控",
        "🔵 保留觀察（CPL偏高但有轉換潛力）": "保留觀察",
        "🟠 CPL過高，評估是否優化素材": "評估優化",
        "❌ 立即關閉":               "立即關閉",
        "❌ 建議關閉（無諮詢轉換）":  "建議關閉",
        "⚪ 待觀察":                  "待觀察",
    }
    consult_rate = r['n_consult']/r['s_leads'] if r['s_leads'] else None
    vals = [r['ad_name'], r['ad_id'], r['spend_full'], r['leads_full'], r['cpl_full'],
            r['s_leads'] or "—", r['n_consult'], r['cpql'], r['cpql_attended'], consult_rate,
            cpl_label, roi_label, r['quad'], action_map.get(r['final'], r['final'])]
    nfs  = [None,None,"$#,##0.00","#,##0","$#,##0.00","#,##0","#,##0","$#,##0.00","$#,##0.00","0.0%",None,None,None,None]
    base_bg = "EBF3FB" if ri%2==0 else "FFFFFF"
    for ci, (v, nf) in enumerate(zip(vals, nfs), 1):
        if ci == 11:
            wc(ws4, ri, ci, v, bold=True, bg=cpl_label_bg, nf=nf, color=cpl_label_fg, align=no_wrap("center"))
        elif ci == 12:
            roi_bg = "D4EDDA" if roi_high else "F8D7DA"
            roi_fg = "1E7E34" if roi_high else "721C24"
            wc(ws4, ri, ci, v, bold=True, bg=roi_bg, nf=nf, color=roi_fg, align=no_wrap("center"))
        elif ci == 13:
            wc(ws4, ri, ci, v, bold=True, bg=qbg, nf=nf, color=qfg, align=center())
        elif ci == 14:
            fbg, ffg = FINAL_C.get(r['final'],("FFFFFF","000000"))
            wc(ws4, ri, ci, v, bold=True, bg=fbg, nf=nf, color=ffg, align=no_wrap("center"))
        else:
            al = no_wrap("center") if ci in {3,4,5,6,7,8,9,10} else left()
            wc(ws4, ri, ci, v, bg=base_bg, nf=nf, align=al)

# PART 2：趨勢矩陣
part2_row = len(rows) + 5
section_title(ws4, part2_row, "▌ PART 2｜4天 / 7天 CPL 趨勢矩陣", 15)
h4b = ["廣告名稱","廣告編號","本期CPL","7天CPL","4天CPL",
       "7天成果數","4天成果數","7天花費\n(USD)","4天花費\n(USD)",
       "7天表現","4天表現","趨勢判斷","趨勢建議動作","四象限","最終綜合建議"]
hdr(ws4, part2_row+1, h4b)

for ri, r in enumerate(rows, part2_row+2):
    ws4.row_dimensions[ri].height = 22
    tbg, tfg = TREND_C.get(r['trend'],("FFFFFF","000000"))
    qbg, qfg = QUAD_C.get(r['quad'],("FFFFFF","000000"))
    fbg, ffg = FINAL_C.get(r['final'],("FFFFFF","000000"))
    base_bg  = "EBF3FB" if ri%2==0 else "FFFFFF"
    t7_str   = {"good":"好","warn":"差","bad":"差"}.get(cpl_tier(r['cpl_7d']),"—") if r['cpl_7d'] else "—"
    t4_str   = {"good":"好","warn":"差","bad":"差"}.get(cpl_tier(r['cpl_4d']),"—") if r['cpl_4d'] else "—"
    vals = [r['ad_name'], r['ad_id'],
            r['cpl_full'], r['cpl_7d'], r['cpl_4d'],
            r['leads_7d'], r['leads_4d'], r['spend_7d'], r['spend_4d'],
            t7_str, t4_str, r['trend'], r['trend_act'], r['quad'], r['final']]
    nfs  = [None,None,"$#,##0.00","$#,##0.00","$#,##0.00",
            "#,##0","#,##0","$#,##0.00","$#,##0.00",None,None,None,None,None,None]
    for ci, (v, nf) in enumerate(zip(vals, nfs), 1):
        if ci == 12:
            wc(ws4, ri, ci, v, bold=True, bg=tbg, nf=nf, color=tfg, align=center())
        elif ci == 13:
            wc(ws4, ri, ci, v, bold=True, bg=tbg, nf=nf, color=tfg, align=center())
        elif ci == 14:
            wc(ws4, ri, ci, v, bold=True, bg=qbg, nf=nf, color=qfg, align=center())
        elif ci == 15:
            wc(ws4, ri, ci, v, bold=True, bg=fbg, nf=nf, color=ffg, align=center())
        else:
            wc(ws4, ri, ci, v, bg=base_bg, nf=nf)

for ci, w in enumerate([28,18,11,10,10,9,8,11,11,9,7,7,18,12,16,26], 1): cw(ws4, ci, w)
ws4.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6：📋 決策摘要
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("📋 決策摘要")
ws5.sheet_view.showGridLines = False
ws5.merge_cells("A1:F1")
ws5["A1"].value = f"廣告決策行動清單｜優先順序排列｜{args.period}"
ws5["A1"].font  = Font(name=FONT, bold=True, size=13, color="FFFFFF")
ws5["A1"].fill  = fl("1A2E44"); ws5["A1"].alignment = center()
ws5.row_dimensions[1].height = 34

decision_groups = [
    ("❌ 立即關閉",              ["❌ 立即關閉"],                              "F8D7DA","721C24"),
    ("❌ 建議關閉",              ["❌ 建議關閉（無諮詢轉換）"],                 "F8D7DA","721C24"),
    ("✅ 強烈建議加碼",          ["✅ 強烈建議加碼"],                          "D4EDDA","155724"),
    ("👀 優質素材（持續監控）",  ["👀 優質素材，持續監控趨勢"],               "D1ECF1","0C5460"),
    ("⚠️ 特殊案例（CPL惡化）",  ["⚠️ 曾帶來諮詢但CPL惡化，建議暫停優化"],   "FFF3CD","856404"),
    ("⚠️ 謹慎觀察",             ["⚠️ 謹慎保留，觀察是否帶出諮詢",
                                  "🟡 CPL偏高且無轉換，密切監控",
                                  "🔵 保留觀察（CPL偏高但有轉換潛力）",
                                  "🟠 CPL過高，評估是否優化素材"],            "FFF3CD","856404"),
    ("⚪ 待觀察",                ["⚪ 待觀察"],                                "F8F9FA","6C757D"),
]

cur_row = 2
col_hdrs = ["廣告名稱","本期CPL","CPQL","7天CPL","4天CPL","建議說明"]

for group_label, final_keys, bg, fg in decision_groups:
    group_rows = [r for r in rows if r['final'] in final_keys]
    if not group_rows:
        continue
    section_title(ws5, cur_row, group_label, 6, bg=bg)
    ws5.cell(cur_row, 1).font = Font(name=FONT, bold=True, size=11, color=fg)
    hdr(ws5, cur_row+1, col_hdrs, bg="34495E")
    for r in group_rows:
        cur_row += 2
        ws5.row_dimensions[cur_row].height = 22
        # Build description
        parts = []
        if r['cpl_full']:  parts.append(f"CPL ${r['cpl_full']:.2f}")
        if r['cpl_7d']:    parts.append(f"7天 ${r['cpl_7d']:.2f}")
        if r['cpl_4d']:    parts.append(f"4天 ${r['cpl_4d']:.2f}")
        if r['n_consult']: parts.append(f"帶來 {r['n_consult']} 筆諮詢（{r['consult_names']}）")
        desc = "｜".join(parts)
        vals = [r['ad_name'], r['cpl_full'], r['cpql'], r['cpl_7d'], r['cpl_4d'], desc]
        nfs  = [None,"$#,##0.00","$#,##0.00","$#,##0.00","$#,##0.00",None]
        for ci, (v, nf) in enumerate(zip(vals, nfs), 1):
            al = no_wrap("center") if ci in {2,3,4,5} else left()
            wc(ws5, cur_row, ci, v, bg=bg, nf=nf, align=al,
               color=fg if ci == 1 else "000000",
               bold=(ci==1))
    cur_row += 2

for ci, w in enumerate([30,11,11,11,8,46], 1): cw(ws5, ci, w)
ws5.freeze_panes = "A3"

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(args.output)
print(json.dumps({
    "output": args.output,
    "sheets": ["📊 執行摘要","📢 廣告素材效益","📋 諮詢比對明細","📥 Systeme名單統計","🎯 廣告決策矩陣","📋 決策摘要"],
    "total_ads": len(rows),
    "total_brand_consults": len(brand_consults),
    "stage2_alerts": len(stage2_alerts),
    "summary": {
        "強烈加碼": sum(1 for r in rows if r['final']=="✅ 強烈建議加碼"),
        "優質監控": sum(1 for r in rows if r['final']=="👀 優質素材，持續監控趨勢"),
        "立即關閉": sum(1 for r in rows if r['final']=="❌ 立即關閉"),
        "建議關閉": sum(1 for r in rows if "建議關閉" in r['final']),
        "特殊案例": sum(1 for r in rows if "曾帶來諮詢" in r['final']),
        "謹慎觀察": sum(1 for r in rows if any(k in r['final'] for k in ["謹慎","警示","偏高","過高"])),
    }
}, ensure_ascii=False, indent=2))
